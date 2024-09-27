import os
from sklearn.preprocessing import StandardScaler
import pandas as pd
import numpy as np
from sklearn import svm
import itertools
import openpyxl


def find_peak(path):
    peak = None
    for f in os.listdir(path):
        if f.__contains__('_p'):
            peak = int(f.split('.')[0].split('_')[0])
    if peak is None:
        raise ValueError('No peak found!')
    else:
        return peak


def handle_us_ceus():
    label_file = "./breast/US_CEUS_feats_short_label.xlsx"
    data_dir = "./breast/feats"
    source_dir = './breast/source_imaging'
    dst_dir = './breast/norm_feats_metastasis'
    if not os.path.exists(dst_dir):
        os.mkdir(dst_dir)
    label_data = pd.read_excel(label_file, sheet_name='Sheet1')
    label_data = label_data.values
    id_label = label_data[:, 1:3]

    us_data = pd.read_excel(os.path.join(data_dir, 'US_ALL.xls'), sheet_name='Sheet1')
    us_data = us_data.values
    _, us_index, _ = np.intersect1d(us_data[:, 0:1], id_label[:, 0:1], return_indices=True)

    select_us_data = np.hstack([id_label, us_data[us_index, 1:]])
    ceus_peak_datas = []
    for i in select_us_data:
        image_path = os.path.join(source_dir, str(int(i[0])), 'imgs')
        if not os.path.exists(image_path):
            image_path = os.path.join(source_dir, str(int(i[0])) + '良', 'imgs')
        peak = find_peak(image_path)
        ceus_case_data = pd.read_excel(os.path.join(data_dir, f'CEUS_{str(int(i[0]))}.xls'), sheet_name='Sheet1')
        ceus_case_data = ceus_case_data.values
        ceus_peak_datas.append(np.squeeze(ceus_case_data[ceus_case_data[:, 1] == peak]))
    for file in os.listdir(data_dir):
        index = os.path.splitext(file)[0].split('_')[1]
        if not index.isdigit():
            continue
        image_path = os.path.join(source_dir, str(int(index)), 'imgs')
        if not os.path.exists(image_path):
            image_path = os.path.join(source_dir, str(int(index)) + '良', 'imgs')
        peak = find_peak(image_path)
        ceus_case_data = pd.read_excel(os.path.join(data_dir, f'CEUS_{str(int(i[0]))}.xls'), sheet_name='Sheet1')
        ceus_case_data = ceus_case_data.values
        ceus_peak_datas.append(np.squeeze(ceus_case_data[ceus_case_data[:, 1] == peak]))
        dst_file_path = os.path.join(dst_dir, f'CEUS_{str(int(index))}.xlsx')
        idx = np.where(ceus_case_data[:, 1] == peak)[0][0] + 1
        pd.DataFrame(ceus_case_data[:, 2:]).to_excel(dst_file_path, sheet_name='feats', float_format='%.5f')
        workbook = openpyxl.load_workbook(dst_file_path)
        sheet2 = workbook.create_sheet('idx', index=2)
        sheet2['A1'] = idx
        workbook.save(dst_file_path)

    ceus_peak_datas = np.array(ceus_peak_datas)
    _, ceus_index, _ = np.intersect1d(ceus_peak_datas[:, 0:1], id_label[:, 0:1], return_indices=True)
    select_data = np.hstack([select_us_data, ceus_peak_datas[ceus_index, 1:]])
    excel_file_path = './breast/US_CEUS_feats_short.xlsx'
    pd.DataFrame(select_data).to_excel(excel_file_path, float_format='%.5f')


if __name__ == "__main__":
    handle_us_ceus()
    exit(0)
    X = os.listdir('./breast/source_imaging')
    y = {}
    for x in X:
        if x.__contains__('良'):
            y[x[:-1]] = 0
        else:
            y[x] = 1

    data_path = "./breast/feats"
    scaler_ceus = StandardScaler()
    scaler_us = StandardScaler()
    ceus_data = []
    ceus_data_dict = {}
    us_data = None
    us_ceus_data = []
    for f in os.listdir(data_path):
        if f.endswith('.xls'):
            if f.startswith('CEUS'):
                xls_file = os.path.join(data_path, f)
                sheet = pd.read_excel(xls_file, sheet_name='Sheet1')
                ceus_data.append(sheet.values)
                sample_id = f.split('.')[0].split('_')[1]
                if os.path.exists(os.path.join('./breast/source_imaging', sample_id)):
                    peak = find_peak(os.path.join('./breast/source_imaging', sample_id, 'imgs'))
                else:
                    peak = find_peak(os.path.join('./breast/source_imaging', sample_id + '良', 'imgs'))
                ceus_data_dict[f.split('.')[0].split('_')[1]] = [peak, sheet.values]
            if f.startswith('US'):
                xls_file = os.path.join(data_path, f)
                sheet = pd.read_excel(xls_file, sheet_name='Sheet1')
                us_data = sheet.values

    ceus_data = np.concatenate(ceus_data, axis=0)
    ceus_data = ceus_data[:, 2:]
    scaler_ceus.fit(ceus_data)
    scaler_us.fit(us_data[:, 1:])
    rows = us_data.shape[0]

    for i in range(rows):
        sample_id = str(int(us_data[i, 0]))
        if sample_id in ceus_data_dict.keys():
            peak = ceus_data_dict[sample_id][0]
            ceus_feat = ceus_data_dict[sample_id][1]
            us_feat = us_data[i, 1:]
            ceus_tid = ceus_feat[:, 1]
            ceus_tid = ceus_tid.astype(int)
            # normalize
            us_ceus_feat = np.concatenate((
                scaler_us.transform(us_feat.reshape(1, -1)),
                scaler_ceus.transform(ceus_feat[ceus_tid == peak, 2:])),
                axis=1)
            us_ceus_feat = np.insert(us_ceus_feat, 0, float(sample_id))
            us_ceus_feat = np.insert(us_ceus_feat, 1, float(y[sample_id]))
            us_ceus_data.append(us_ceus_feat)
            ceus_feat_norm = scaler_ceus.transform(ceus_feat[:, 2:])
            idx = np.where(ceus_tid == peak)[0][0] + 1
            excel_file_path = './breast/norm_feats/CEUS_' + sample_id + '.xlsx'
            # pd.DataFrame(ceus_feat_norm).to_excel(excel_file_path, sheet_name='feats', float_format='%.5f')

            workbook = openpyxl.load_workbook(excel_file_path)
            sheet2 = workbook.create_sheet('idx', index=2)
            sheet2['A1'] = idx
            # 保存修改后的 Excel 文件
            workbook.save(excel_file_path)
        else:
            raise ValueError('No sample_id found!')

    # pd.DataFrame(us_ceus_data).to_excel('E:/data/breast/US_CEUS_feats_short.xlsx', float_format='%.5f')
